from openpyxl import load_workbook
import os, sys
web_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../"))
sys.path.append(web_path)

# class Excel:
#     @staticmethod
#     def get_login_data():
workbook = load_workbook(os.path.join(web_path, "resources", "testdata", "test_data.xlsx"))
sheet = workbook["LoginData"]
data = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    email, password = row[:2]
    data.append((email, password))
    print(row)
# return data
email, password = data
print("aaa",email, password )
